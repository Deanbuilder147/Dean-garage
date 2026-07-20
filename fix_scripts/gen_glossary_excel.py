#!/usr/bin/env python3
"""导出当前词条库为 Excel（按 skillContract 模板）。
输出: ~/Desktop/词条库_Excel模板_2026-07-21.xlsx
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "/Users/dingxuyang/CodeBuddy/20260604120036/mecha-universe-engine/services/combat-service/src/config/glossary-skill-config.json"
OUT = os.path.expanduser("~/Desktop/词条库_Excel模板_2026-07-21.xlsx")

with open(SRC, "r", encoding="utf-8") as f:
    cfg = json.load(f)

# ── 样式 ──
hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
note_fill = PatternFill("solid", fgColor="FFF2CC")
note_font = Font(italic=True, color="7F6000", size=10)
wrap = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def write_rows(ws, headers, rows, widths=None):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
            cell.border = border

wb = Workbook()

# ============ Sheet 1: skills（主表，每行一个词条） ============
MAPPED = {
    "type", "label", "category", "description", "deterministic", "reduction",
    "trigger", "target_filter", "cast_range", "aoe_radius", "base_damage",
    "status_effects", "damage_kind", "min_cast_range", "accuracy_mod",
    "evasion_mod", "height_bonus_per_diff", "action_type", "attack_stat",
    "requires_unmoved", "requires_stealth",
}
SKILL_HEADERS = [
    "key", "name", "category", "target_scope", "cast_range_min", "cast_range_max",
    "skill_shape", "damage_kind", "base_damage", "action_type", "type",
    "description", "deterministic", "status_effects", "accuracy_mod", "evasion_mod",
    "height_bonus_per_diff", "attack_stat", "requires_unmoved", "requires_stealth",
    "has_dice", "dice_type", "dice_branches", "extra_params",
]
skill_rows = []
for key, s in cfg.get("skills", {}).items():
    extra = {k: v for k, v in s.items() if k not in MAPPED}
    status = s.get("status_effects", [])
    status_str = ",".join(status) if isinstance(status, list) else str(status)
    base = [
        key,
        s.get("label", ""),
        s.get("category", ""),
        s.get("target_filter", ""),          # 旧字段，读取时映射到 target_scope
        s.get("min_cast_range", ""),
        s.get("cast_range", ""),
        "",                                   # 旧数据无 skill_shape，默认 single
        s.get("damage_kind", ""),
        s.get("base_damage", ""),
        s.get("action_type", ""),
        s.get("type", ""),
        s.get("description", ""),
        s.get("deterministic", ""),
        status_str,
        s.get("accuracy_mod", ""),
        s.get("evasion_mod", ""),
        s.get("height_bonus_per_diff", ""),
        s.get("attack_stat", ""),
        s.get("requires_unmoved", ""),
        s.get("requires_stealth", ""),
        "false",                              # 当前无投骰多判定
        "6",
        "",                                   # dice_branches JSON（空）
        json.dumps(extra, ensure_ascii=False) if extra else "",
    ]
    skill_rows.append(base)

ws = wb.active
ws.title = "skills"
write_rows(ws, SKILL_HEADERS, skill_rows,
           widths=[12, 12, 10, 12, 12, 12, 10, 11, 11, 11, 8, 40, 11, 12, 11, 11, 14, 11, 14, 14, 9, 9, 30, 40])

# ============ Sheet 2: branches（投骰分支子表，当前为空） ============
BRANCH_HEADERS = ["skill_key", "branch_index", "points", "effects"]
wsb = wb.create_sheet("branches")
write_rows(wsb, BRANCH_HEADERS, [], widths=[14, 12, 14, 40])
# 说明行（第2行起为数据，此说明写在第1行下方用注释色，不占数据行）
note = wsb.cell(row=2, column=1,
                value="（当前词条库无投骰多判定词条，此表为空）每行一个判定分支：skill_key 关联 skills!key；"
                      "points 填 6(精确) 或 4-8(区间)；effects 填 动作:数值@目标，如 damage:10;heal:8@ally")
note.fill = note_fill
note.font = note_font
note.alignment = wrap
wsb.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

# ============ Sheet 3: terrains ============
TERRAIN_HEADERS = ["key", "name", "color", "move_cost", "defense_bonus",
                   "is_destructible", "max_hp", "destroyed_transform_to", "damage_kind_modifiers"]
terrain_rows = []
for k, t in cfg.get("terrains", {}).items():
    dkm = t.get("damage_kind_modifiers", {})
    terrain_rows.append([
        k, t.get("name", ""), t.get("color", ""), t.get("move_cost", ""),
        t.get("defense_bonus", ""), t.get("is_destructible", ""), t.get("max_hp", ""),
        t.get("destroyed_transform_to", ""),
        json.dumps(dkm, ensure_ascii=False),
    ])
wt = wb.create_sheet("terrains")
write_rows(wt, TERRAIN_HEADERS, terrain_rows,
           widths=[14, 12, 10, 10, 13, 13, 8, 20, 50])

# ============ Sheet 4: damage_kinds ============
DK_HEADERS = ["key", "label", "description"]
dk_rows = [[k, v.get("label", ""), v.get("description", "")] for k, v in cfg.get("damage_kinds", {}).items()]
wd = wb.create_sheet("damage_kinds")
write_rows(wd, DK_HEADERS, dk_rows, widths=[14, 12, 40])

# ============ Sheet 5: action_types ============
AT_HEADERS = ["key", "label", "description"]
at_rows = [[k, v.get("label", ""), v.get("description", "")] for k, v in cfg.get("action_types", {}).items()]
wa = wb.create_sheet("action_types")
write_rows(wa, AT_HEADERS, at_rows, widths=[14, 12, 40])

# ============ Sheet 6: systems ============
SYS_HEADERS = ["key", "label", "description", "extra_params"]
sys_rows = []
for k, v in cfg.get("systems", {}).items():
    e = {kk: vv for kk, vv in v.items() if kk not in ("label", "description")}
    sys_rows.append([k, v.get("label", ""), v.get("description", ""),
                     json.dumps(e, ensure_ascii=False) if e else ""])
wsy = wb.create_sheet("systems")
write_rows(wsy, SYS_HEADERS, sys_rows, widths=[14, 12, 50, 50])

# ============ Sheet 7: README（模板说明） ============
wr = wb.create_sheet("README")
readme_lines = [
    ["词条库 Excel 模板说明（与 skillContract.js 1:1 对齐）"],
    [""],
    ["【skills 主表】每行一个词条，更新时按 key 深度合并(upsert)，key 一致即覆盖。"],
    ["字段约定："],
    ["  target_scope  填旧字段 target_filter 值(self/enemy/ally/all)，读取时自动映射 enemy/ally 等"],
    ["  cast_range   用 cast_range_min / cast_range_max 两列（或旧字段 min_cast_range / cast_range）"],
    ["  skill_shape  枚举: single / fan / linear / concentric（旧库留空则默认 single）"],
    ["  category     枚举: melee / ranged / automation / support（旧库 special 读取时归为 melee）"],
    ["  damage_kind  枚举: kinetic / beam / explosive / corrosive / thermal（energy/laser/em 自动归一 beam）"],
    ["  action_type  枚举: attack / heal / buff / debuff / passive"],
    ["  type         枚举: active / passive"],
    ["  attack_stat  枚举: melee / ranged / max"],
    ["  status_effects  多个用逗号分隔，如 burn,stun"],
    ["  deterministic / requires_unmoved / requires_stealth  填 true / false"],
    ["  extra_params 旧版专属参数(reduction/trigger/sector_angle/mode/effect/value 等)以 JSON 保留，回传时合并"],
    [""],
    ["【branches 子表】投骰多判定词条用：skill_key 关联主表 key；"],
    ["  points 填 6(精确) 或 4-8(区间)；effects 填 动作:数值@目标，如 damage:10;heal:8@ally"],
    ["  动作枚举: damage / damage_bonus / heal / apply_status / mobility_mod / accuracy_mod"],
    [""],
    ["【dice_branches 列】若不用 branches 子表，可在 skills 表 dice_branches 列直接填 JSON："],
    ['  [{"points":[{"kind":"exact","value":6}],"effects":[{"action":"damage","value":10}]}]'],
    [""],
    ["【其余分区】terrains / damage_kinds / action_types / systems 为独立表，结构见各表头。"],
    ["保存后可将此文件交给「词条 Excel 读取功能」上传，自动 upsert 进 glossary-skill-config.json。"],
]
for r in readme_lines:
    wr.append(r)
wr.column_dimensions["A"].width = 110
for row in wr.iter_rows():
    for cell in row:
        cell.alignment = wrap
wr["A1"].font = Font(bold=True, size=13, color="2F5496")

# 把 README 放到最前便于查看
wb.move_sheet("README", -(len(wb.sheetnames) - 1))

wb.save(OUT)
print("已生成:", OUT)
print("sheets:", wb.sheetnames)
print("skills 行数:", len(skill_rows))
